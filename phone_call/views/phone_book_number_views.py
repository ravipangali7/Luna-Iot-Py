"""
Phone Book Number Views
Handles phone book number management endpoints
"""
from rest_framework.decorators import api_view
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from io import BytesIO
from openpyxl import Workbook
from phone_call.models import PhoneBookNumber, PhoneBook
from phone_call.serializers import (
    PhoneBookNumberSerializer,
    PhoneBookNumberCreateSerializer,
    PhoneBookNumberUpdateSerializer,
    PhoneBookNumberListSerializer
)
from phone_call.services.phone_book_number_importer import PhoneBookNumberImporter
from api_common.utils.response_utils import success_response, error_response
from api_common.constants.api_constants import SUCCESS_MESSAGES, ERROR_MESSAGES, HTTP_STATUS
from api_common.decorators.response_decorators import api_response
from api_common.decorators.auth_decorators import require_auth
from api_common.exceptions.api_exceptions import NotFoundError
from core.models import InstituteModule


def require_phone_book_number_access(phone_book_id_param='phone_book_id'):
    """
    Decorator to require access to phone book before managing its numbers
    
    Args:
        phone_book_id_param: Name of the URL parameter containing the phone book ID
    """
    def decorator(view_func):
        from functools import wraps
        from api_common.utils.response_utils import error_response
        
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not hasattr(request, 'user') or not request.user.is_authenticated:
                return error_response(
                    message='Authentication required',
                    status_code=401
                )
            
            phone_book_id = kwargs.get(phone_book_id_param)
            
            if phone_book_id:
                try:
                    phone_book = PhoneBook.objects.get(id=phone_book_id)
                except PhoneBook.DoesNotExist:
                    return error_response(
                        message='Phone book not found',
                        status_code=404
                    )
                
                # Check if user is Super Admin - always allow
                user_groups = request.user.groups.all()
                user_role_names = [group.name for group in user_groups]
                is_super_admin = 'Super Admin' in user_role_names
                
                if not is_super_admin:
                    # Check ownership
                    is_owner = False
                    if phone_book.user and phone_book.user.id == request.user.id:
                        is_owner = True
                    elif phone_book.institute:
                        has_access = InstituteModule.objects.filter(
                            institute=phone_book.institute,
                            users=request.user
                        ).exists()
                        if has_access:
                            is_owner = True
                    
                    if not is_owner:
                        return error_response(
                            message='Access denied. You do not have permission to manage this phone book',
                            status_code=403
                        )
            else:
                # For operations without phone_book_id, check from request data
                if request.method == 'POST':
                    phone_book_id_from_data = request.data.get('phonebook')
                    if phone_book_id_from_data:
                        # Handle different formats
                        if isinstance(phone_book_id_from_data, (int, float)):
                            phone_book_id_from_data = int(phone_book_id_from_data)
                        elif isinstance(phone_book_id_from_data, dict):
                            phone_book_id_from_data = phone_book_id_from_data.get('id') or phone_book_id_from_data.get('pk')
                        elif hasattr(phone_book_id_from_data, 'id'):
                            phone_book_id_from_data = phone_book_id_from_data.id
                        elif isinstance(phone_book_id_from_data, str) and phone_book_id_from_data.isdigit():
                            phone_book_id_from_data = int(phone_book_id_from_data)
                        
                        try:
                            phone_book = PhoneBook.objects.get(id=phone_book_id_from_data)
                        except PhoneBook.DoesNotExist:
                            return error_response(
                                message='Phone book not found',
                                status_code=404
                            )
                        
                        # Check if user is Super Admin - always allow
                        user_groups = request.user.groups.all()
                        user_role_names = [group.name for group in user_groups]
                        is_super_admin = 'Super Admin' in user_role_names
                        
                        if not is_super_admin:
                            # Check ownership
                            is_owner = False
                            if phone_book.user and phone_book.user.id == request.user.id:
                                is_owner = True
                            elif phone_book.institute:
                                has_access = InstituteModule.objects.filter(
                                    institute=phone_book.institute,
                                    users=request.user
                                ).exists()
                                if has_access:
                                    is_owner = True
                            
                            if not is_owner:
                                return error_response(
                                    message='Access denied. You do not have permission to manage this phone book',
                                    status_code=403
                                )
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


@api_view(['GET'])
@require_auth
@require_phone_book_number_access('phone_book_id')
@api_response
def get_phone_book_numbers(request, phone_book_id):
    """Get all numbers for a phone book"""
    try:
        try:
            phone_book = PhoneBook.objects.get(id=phone_book_id)
        except PhoneBook.DoesNotExist:
            raise NotFoundError("Phone book not found")
        
        numbers = PhoneBookNumber.objects.filter(phonebook=phone_book).order_by('name')
        serializer = PhoneBookNumberListSerializer(numbers, many=True)
        
        return success_response(
            data=serializer.data,
            message=SUCCESS_MESSAGES.get('DATA_RETRIEVED', 'Phone book numbers retrieved successfully')
        )
    except NotFoundError as e:
        return error_response(
            message=str(e),
            status_code=HTTP_STATUS['NOT_FOUND']
        )
    except Exception as e:
        return error_response(
            message=ERROR_MESSAGES.get('INTERNAL_ERROR', 'Internal server error'),
            data=str(e)
        )


@api_view(['GET'])
@require_auth
@api_response
def get_phone_book_number_by_id(request, phone_book_id, number_id):
    """Get phone book number by ID"""
    try:
        try:
            phone_book = PhoneBook.objects.get(id=phone_book_id)
        except PhoneBook.DoesNotExist:
            raise NotFoundError("Phone book not found")
        
        # Check access to phone book
        user_groups = request.user.groups.all()
        user_role_names = [group.name for group in user_groups]
        is_super_admin = 'Super Admin' in user_role_names
        
        if not is_super_admin:
            is_owner = False
            if phone_book.user and phone_book.user.id == request.user.id:
                is_owner = True
            elif phone_book.institute:
                has_access = InstituteModule.objects.filter(
                    institute=phone_book.institute,
                    users=request.user
                ).exists()
                if has_access:
                    is_owner = True
            
            if not is_owner:
                return error_response(
                    message='Access denied. You do not have permission to view this phone book',
                    status_code=403
                )
        
        try:
            number = PhoneBookNumber.objects.get(id=number_id, phonebook=phone_book)
        except PhoneBookNumber.DoesNotExist:
            raise NotFoundError("Phone book number not found")
        
        serializer = PhoneBookNumberSerializer(number)
        
        return success_response(
            data=serializer.data,
            message=SUCCESS_MESSAGES.get('DATA_RETRIEVED', 'Phone book number retrieved successfully')
        )
    except NotFoundError as e:
        return error_response(
            message=str(e),
            status_code=HTTP_STATUS['NOT_FOUND']
        )
    except Exception as e:
        return error_response(
            message=ERROR_MESSAGES.get('INTERNAL_ERROR', 'Internal server error'),
            data=str(e)
        )


@api_view(['POST'])
@require_auth
@require_phone_book_number_access('phone_book_id')
@api_response
def create_phone_book_number(request, phone_book_id):
    """Create new phone book number"""
    try:
        try:
            phone_book = PhoneBook.objects.get(id=phone_book_id)
        except PhoneBook.DoesNotExist:
            raise NotFoundError("Phone book not found")
        
        # Add phonebook to request data if not present
        data = request.data.copy()
        if 'phonebook' not in data:
            data['phonebook'] = phone_book_id
        
        serializer = PhoneBookNumberCreateSerializer(data=data)
        
        if serializer.is_valid():
            number = serializer.save()
            response_serializer = PhoneBookNumberSerializer(number)
            
            return success_response(
                data=response_serializer.data,
                message=SUCCESS_MESSAGES.get('DATA_CREATED', 'Phone book number created successfully'),
                status_code=HTTP_STATUS['CREATED']
            )
        else:
            return error_response(
                message=ERROR_MESSAGES.get('VALIDATION_ERROR', 'Validation error'),
                data=serializer.errors,
                status_code=HTTP_STATUS['BAD_REQUEST']
            )
    except NotFoundError as e:
        return error_response(
            message=str(e),
            status_code=HTTP_STATUS['NOT_FOUND']
        )
    except Exception as e:
        return error_response(
            message=ERROR_MESSAGES.get('INTERNAL_ERROR', 'Internal server error'),
            data=str(e)
        )


@api_view(['PUT'])
@require_auth
@require_phone_book_number_access('phone_book_id')
@api_response
def update_phone_book_number(request, phone_book_id, number_id):
    """Update phone book number"""
    try:
        try:
            phone_book = PhoneBook.objects.get(id=phone_book_id)
        except PhoneBook.DoesNotExist:
            raise NotFoundError("Phone book not found")
        
        try:
            number = PhoneBookNumber.objects.get(id=number_id, phonebook=phone_book)
        except PhoneBookNumber.DoesNotExist:
            raise NotFoundError("Phone book number not found")
        
        serializer = PhoneBookNumberUpdateSerializer(number, data=request.data)
        
        if serializer.is_valid():
            number = serializer.save()
            response_serializer = PhoneBookNumberSerializer(number)
            
            return success_response(
                data=response_serializer.data,
                message=SUCCESS_MESSAGES.get('DATA_UPDATED', 'Phone book number updated successfully')
            )
        else:
            return error_response(
                message=ERROR_MESSAGES.get('VALIDATION_ERROR', 'Validation error'),
                data=serializer.errors,
                status_code=HTTP_STATUS['BAD_REQUEST']
            )
    except NotFoundError as e:
        return error_response(
            message=str(e),
            status_code=HTTP_STATUS['NOT_FOUND']
        )
    except Exception as e:
        return error_response(
            message=ERROR_MESSAGES.get('INTERNAL_ERROR', 'Internal server error'),
            data=str(e)
        )


@api_view(['DELETE'])
@require_auth
@require_phone_book_number_access('phone_book_id')
@api_response
def delete_phone_book_number(request, phone_book_id, number_id):
    """Delete phone book number"""
    try:
        try:
            phone_book = PhoneBook.objects.get(id=phone_book_id)
        except PhoneBook.DoesNotExist:
            raise NotFoundError("Phone book not found")
        
        try:
            number = PhoneBookNumber.objects.get(id=number_id, phonebook=phone_book)
        except PhoneBookNumber.DoesNotExist:
            raise NotFoundError("Phone book number not found")
        
        number_name = number.name
        number.delete()
        
        return success_response(
            data={'id': number_id},
            message=f"Phone book number '{number_name}' deleted successfully"
        )
    except NotFoundError as e:
        return error_response(
            message=str(e),
            status_code=HTTP_STATUS['NOT_FOUND']
        )
    except Exception as e:
        return error_response(
            message=ERROR_MESSAGES.get('INTERNAL_ERROR', 'Internal server error'),
            data=str(e)
        )


@api_view(['POST'])
@require_auth
@require_phone_book_number_access('phone_book_id')
@api_response
def bulk_create_phone_book_numbers(request, phone_book_id):
    """Bulk create phone book numbers"""
    try:
        try:
            phone_book = PhoneBook.objects.get(id=phone_book_id)
        except PhoneBook.DoesNotExist:
            raise NotFoundError("Phone book not found")
        
        numbers_data = request.data.get('numbers', [])
        
        if not isinstance(numbers_data, list):
            return error_response(
                message='Numbers must be a list',
                status_code=HTTP_STATUS['BAD_REQUEST']
            )
        
        if not numbers_data:
            return error_response(
                message='At least one number is required',
                status_code=HTTP_STATUS['BAD_REQUEST']
            )
        
        created_numbers = []
        errors = []
        
        for idx, number_data in enumerate(numbers_data):
            number_data['phonebook'] = phone_book_id
            serializer = PhoneBookNumberCreateSerializer(data=number_data)
            
            if serializer.is_valid():
                number = serializer.save()
                created_numbers.append(PhoneBookNumberSerializer(number).data)
            else:
                errors.append({
                    'index': idx,
                    'data': number_data,
                    'errors': serializer.errors
                })
        
        return success_response(
            data={
                'created': created_numbers,
                'errors': errors,
                'created_count': len(created_numbers),
                'error_count': len(errors)
            },
            message=f"Bulk create completed. {len(created_numbers)} created, {len(errors)} errors"
        )
    except NotFoundError as e:
        return error_response(
            message=str(e),
            status_code=HTTP_STATUS['NOT_FOUND']
        )
    except Exception as e:
        return error_response(
            message=ERROR_MESSAGES.get('INTERNAL_ERROR', 'Internal server error'),
            data=str(e)
        )


@csrf_exempt
@require_http_methods(["POST"])
@require_auth
@require_phone_book_number_access('phone_book_id')
@api_response
def upload_phone_book_numbers_excel(request, phone_book_id):
    """Upload Excel/CSV file to import phone book numbers"""
    try:
        try:
            phone_book = PhoneBook.objects.get(id=phone_book_id)
        except PhoneBook.DoesNotExist:
            raise NotFoundError("Phone book not found")
        
        if 'file' not in request.FILES:
            return error_response(
                message='No file provided',
                status_code=HTTP_STATUS['BAD_REQUEST']
            )
        
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        
        # Determine file type
        if file_name.endswith('.csv'):
            file_type = 'csv'
        elif file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            file_type = 'xlsx'
        else:
            return error_response(
                message='Unsupported file type. Please upload CSV or XLSX file.',
                status_code=HTTP_STATUS['BAD_REQUEST']
            )
        
        # Validate file size (max 10MB)
        if uploaded_file.size > 10 * 1024 * 1024:
            return error_response(
                message='File size exceeds 10MB limit',
                status_code=HTTP_STATUS['BAD_REQUEST']
            )
        
        # Import data
        importer = PhoneBookNumberImporter()
        result = importer.import_phone_book_numbers(uploaded_file, phone_book_id, file_type)
        
        if result.get('success', False):
            return success_response(
                data=result,
                message=f"Import completed. {result.get('successful', 0)} contacts imported, {result.get('failed', 0)} failed"
            )
        else:
            return error_response(
                result.get('error', 'Import failed'),
                HTTP_STATUS['BAD_REQUEST'],
                data=result
            )
    
    except NotFoundError as e:
        return error_response(
            message=str(e),
            status_code=HTTP_STATUS['NOT_FOUND']
        )
    except Exception as e:
        return error_response(
            message=ERROR_MESSAGES.get('INTERNAL_ERROR', 'Internal server error'),
            data=str(e)
        )


@csrf_exempt
@require_http_methods(["GET"])
@require_auth
@require_phone_book_number_access('phone_book_id')
def download_phone_book_template(request, phone_book_id):
    """Download Excel template file for phone book numbers"""
    try:
        try:
            phone_book = PhoneBook.objects.get(id=phone_book_id)
        except PhoneBook.DoesNotExist:
            return error_response(
                message='Phone book not found',
                status_code=HTTP_STATUS['NOT_FOUND']
            )
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Phone Book Template"
        
        # Add headers
        ws['A1'] = 'Name'
        ws['B1'] = 'Phone'
        
        # Make headers bold
        from openpyxl.styles import Font
        header_font = Font(bold=True)
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        
        # Add sample rows
        ws['A2'] = 'John Doe'
        ws['B2'] = '1234567890'
        ws['A3'] = 'Jane Smith'
        ws['B3'] = '9876543210'
        
        # Create in-memory file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="phone_book_template.xlsx"'
        
        return response
    
    except Exception as e:
        return error_response(
            message=ERROR_MESSAGES.get('INTERNAL_ERROR', 'Internal server error'),
            data=str(e)
        )
